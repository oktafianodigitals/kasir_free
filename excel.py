import openpyxl
import os
import numpy
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from database import Database
from datetime import datetime
from collections import defaultdict
import pandas as pd

class ExcelExporter:
    def __init__(self):
        self.db = Database()
        self.export_folder = 'export'
        self._ensure_export_folder()
    
    def _ensure_export_folder(self):
        """Pastikan folder export ada"""
        if not os.path.exists(self.export_folder):
            os.makedirs(self.export_folder)
    
    def format_rupiah(self, amount):
        """Format angka ke format Rupiah"""
        return f"Rp {amount:,}".replace(',', '.')
    
    def _create_styles(self):
        """Buat style untuk Excel"""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(size=10)
        data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Number style
        number_style = NamedStyle(name="number")
        number_style.font = Font(size=10)
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center style
        center_style = NamedStyle(name="center")
        center_style.font = Font(size=10)
        center_style.alignment = Alignment(horizontal="center", vertical="center")
        center_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(bold=True, size=16, color="2F4F4F")
        title_style.alignment = Alignment(horizontal="center")
        
        # Subtitle style
        subtitle_style = NamedStyle(name="subtitle")
        subtitle_style.font = Font(bold=True, size=12, color="4F4F4F")
        subtitle_style.alignment = Alignment(horizontal="left")
        
        return {
            'header': header_style,
            'data': data_style,
            'number': number_style,
            'center': center_style,
            'title': title_style,
            'subtitle': subtitle_style
        }
    
    def _create_summary_sheet(self, wb, transaksi_list, period_info=""):
        """Buat sheet summary dengan statistik"""
        ws_summary = wb.create_sheet("Summary")
        styles = self._create_styles()
        
        # Title
        title_text = f"LAPORAN SUMMARY TRANSAKSI{' ' + period_info if period_info else ''}"
        ws_summary['A1'] = title_text
        ws_summary['A1'].style = styles['title']
        ws_summary.merge_cells('A1:D1')
        
        # Info export
        ws_summary['A3'] = "Tanggal Export:"
        ws_summary['B3'] = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        
        # Summary calculations
        total_transaksi = len(transaksi_list)
        total_subtotal = sum([t[2] for t in transaksi_list])
        total_pembayaran = sum([t[3] for t in transaksi_list])
        total_lunas = len([t for t in transaksi_list if t[4] == 'LUNAS'])
        total_belum_lunas = total_transaksi - total_lunas
        total_piutang = total_subtotal - total_pembayaran
        
        # Summary data
        summary_data = [
            ("STATISTIK UMUM", ""),
            ("Total Transaksi", total_transaksi),
            ("Total Subtotal", self.format_rupiah(total_subtotal)),
            ("Total Pembayaran", self.format_rupiah(total_pembayaran)),
            ("Total Piutang", self.format_rupiah(total_piutang)),
            ("", ""),
            ("STATUS PEMBAYARAN", ""),
            ("Transaksi Lunas", total_lunas),
            ("Transaksi Belum Lunas", total_belum_lunas),
            ("Persentase Lunas", f"{(total_lunas/total_transaksi*100):.1f}%" if total_transaksi > 0 else "0%"),
        ]
        
        # Tulis summary data
        start_row = 5
        for i, (label, value) in enumerate(summary_data):
            row = start_row + i
            ws_summary[f'A{row}'] = label
            if label and not label.isupper():  # Bukan header
                ws_summary[f'A{row}'].style = styles['data']
                ws_summary[f'B{row}'] = value
                ws_summary[f'B{row}'].style = styles['data']
            else:  # Header section
                ws_summary[f'A{row}'].style = styles['subtitle']
                ws_summary.merge_cells(f'A{row}:B{row}')
        
        # Analisis per bulan (jika ada data)
        monthly_stats = defaultdict(lambda: {'count': 0, 'subtotal': 0, 'pembayaran': 0})
        for transaksi in transaksi_list:
            try:
                tanggal = datetime.strptime(transaksi[5], '%Y-%m-%d')
                month_key = tanggal.strftime('%Y-%m')
                monthly_stats[month_key]['count'] += 1
                monthly_stats[month_key]['subtotal'] += transaksi[2]
                monthly_stats[month_key]['pembayaran'] += transaksi[3]
            except:
                continue
        
        data_row = start_row + len(summary_data)
        
        if monthly_stats:
            # Header analisis bulanan
            month_start_row = start_row + len(summary_data) + 2
            ws_summary[f'A{month_start_row}'] = "ANALISIS PER BULAN"
            ws_summary[f'A{month_start_row}'].style = styles['subtitle']
            ws_summary.merge_cells(f'A{month_start_row}:D{month_start_row}')
            
            # Header tabel
            headers = ['Bulan', 'Jumlah Transaksi', 'Total Subtotal', 'Total Pembayaran']
            header_row = month_start_row + 1
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=header_row, column=col, value=header)
                cell.style = styles['header']
            
            # Data bulanan
            data_row = header_row + 1
            for month, stats in sorted(monthly_stats.items()):
                month_formatted = datetime.strptime(month, '%Y-%m').strftime('%B %Y')
                ws_summary[f'A{data_row}'] = month_formatted
                ws_summary[f'B{data_row}'] = stats['count']
                ws_summary[f'C{data_row}'] = self.format_rupiah(stats['subtotal'])
                ws_summary[f'D{data_row}'] = self.format_rupiah(stats['pembayaran'])
                
                for col in range(1, 5):
                    cell = ws_summary.cell(row=data_row, column=col)
                    cell.style = styles['data']
                    if col in [3, 4]:  # Kolom rupiah
                        cell.style = styles['number']
                    elif col == 2:  # Kolom angka
                        cell.style = styles['center']
                
                data_row += 1
        
        # Analisis per Pembeli
        buyer_stats = defaultdict(lambda: {'count': 0, 'subtotal': 0, 'pembayaran': 0, 'lunas': 0, 'belum_lunas': 0})
        for transaksi in transaksi_list:
            nama_pembeli = transaksi[1]
            buyer_stats[nama_pembeli]['count'] += 1
            buyer_stats[nama_pembeli]['subtotal'] += transaksi[2]
            buyer_stats[nama_pembeli]['pembayaran'] += transaksi[3]
            if transaksi[4] == 'LUNAS':
                buyer_stats[nama_pembeli]['lunas'] += 1
            else:
                buyer_stats[nama_pembeli]['belum_lunas'] += 1
        
        if buyer_stats:
            # Header analisis per pembeli
            buyer_start_row = data_row + 2 if monthly_stats else start_row + len(summary_data) + 2
            ws_summary[f'A{buyer_start_row}'] = "ANALISIS PER PEMBELI"
            ws_summary[f'A{buyer_start_row}'].style = styles['subtitle']
            ws_summary.merge_cells(f'A{buyer_start_row}:E{buyer_start_row}')
            
            # Header tabel
            buyer_headers = ['Nama Pembeli', 'Jumlah Transaksi', 'Total Subtotal', 'Total Pembayaran', 'Total Piutang']
            buyer_header_row = buyer_start_row + 1
            for col, header in enumerate(buyer_headers, 1):
                cell = ws_summary.cell(row=buyer_header_row, column=col, value=header)
                cell.style = styles['header']
            
            # Data per pembeli
            buyer_data_row = buyer_header_row + 1
            for nama_pembeli, stats in sorted(buyer_stats.items()):
                piutang = stats['subtotal'] - stats['pembayaran']
                ws_summary[f'A{buyer_data_row}'] = nama_pembeli
                ws_summary[f'B{buyer_data_row}'] = stats['count']
                ws_summary[f'C{buyer_data_row}'] = self.format_rupiah(stats['subtotal'])
                ws_summary[f'D{buyer_data_row}'] = self.format_rupiah(stats['pembayaran'])
                ws_summary[f'E{buyer_data_row}'] = self.format_rupiah(piutang)
                
                for col in range(1, 6):
                    cell = ws_summary.cell(row=buyer_data_row, column=col)
                    cell.style = styles['data']
                    if col in [3, 4, 5]:  # Kolom rupiah
                        cell.style = styles['number']
                    elif col == 2:  # Kolom angka
                        cell.style = styles['center']
                
                buyer_data_row += 1
            
            # Tabel Status Pembayaran per Pembeli
            status_start_row = buyer_data_row + 2
            ws_summary[f'A{status_start_row}'] = "STATUS PEMBAYARAN PER PEMBELI"
            ws_summary[f'A{status_start_row}'].style = styles['subtitle']
            ws_summary.merge_cells(f'A{status_start_row}:D{status_start_row}')
            
            # Header tabel status
            status_headers = ['Nama Pembeli', 'Transaksi Lunas', 'Transaksi Belum Lunas', 'Total Transaksi']
            status_header_row = status_start_row + 1
            for col, header in enumerate(status_headers, 1):
                cell = ws_summary.cell(row=status_header_row, column=col, value=header)
                cell.style = styles['header']
            
            # Data status per pembeli
            status_data_row = status_header_row + 1
            for nama_pembeli, stats in sorted(buyer_stats.items()):
                ws_summary[f'A{status_data_row}'] = nama_pembeli
                ws_summary[f'B{status_data_row}'] = stats['lunas']
                ws_summary[f'C{status_data_row}'] = stats['belum_lunas']
                ws_summary[f'D{status_data_row}'] = stats['count']
                
                for col in range(1, 5):
                    cell = ws_summary.cell(row=status_data_row, column=col)
                    cell.style = styles['data']
                    if col in [2, 3, 4]:  # Kolom angka
                        cell.style = styles['center']
                
                status_data_row += 1
        
        # Auto-adjust column width
        for col in range(1, 6):
            ws_summary.column_dimensions[get_column_letter(col)].width = 22
        
        return ws_summary
    
    def _create_chart_sheet(self, wb, transaksi_list):
        """Buat sheet dengan chart"""
        ws_chart = wb.create_sheet("Charts")
        styles = self._create_styles()
        
        # Title
        ws_chart['A1'] = "DASHBOARD ANALISIS TRANSAKSI"
        ws_chart['A1'].style = styles['title']
        ws_chart.merge_cells('A1:H1')
        
        # Data untuk chart status pembayaran
        status_count = {'LUNAS': 0, 'BELUM LUNAS': 0}
        for transaksi in transaksi_list:
            status_count[transaksi[4]] += 1
        
        # Tabel data untuk Pie Chart
        chart_start_row = 3
        ws_chart[f'A{chart_start_row}'] = "Status Pembayaran"
        ws_chart[f'A{chart_start_row}'].style = styles['subtitle']
        
        data_headers = ['Status', 'Jumlah']
        for col, header in enumerate(data_headers, 1):
            cell = ws_chart.cell(row=chart_start_row + 1, column=col, value=header)
            cell.style = styles['header']
        
        row = chart_start_row + 2
        for status, count in status_count.items():
            ws_chart[f'A{row}'] = status
            ws_chart[f'B{row}'] = count
            ws_chart[f'A{row}'].style = styles['data']
            ws_chart[f'B{row}'].style = styles['center']
            row += 1
        
        # Buat Pie Chart
        pie_chart = PieChart()
        pie_chart.title = "Distribusi Status Pembayaran"
        pie_chart.height = 10
        pie_chart.width = 15
        
        data = Reference(ws_chart, min_col=2, min_row=chart_start_row + 1, 
                        max_row=chart_start_row + 1 + len(status_count))
        labels = Reference(ws_chart, min_col=1, min_row=chart_start_row + 2, 
                          max_row=chart_start_row + 1 + len(status_count))
        
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        ws_chart.add_chart(pie_chart, f'D{chart_start_row}')
        
        # Data untuk Bar Chart (Transaksi per Bulan)
        monthly_data = defaultdict(int)
        for transaksi in transaksi_list:
            try:
                tanggal = datetime.strptime(transaksi[5], '%Y-%m-%d')
                month_key = tanggal.strftime('%Y-%m')
                monthly_data[month_key] += 1
            except:
                continue
        
        if monthly_data:
            bar_start_row = chart_start_row + 20
            ws_chart[f'A{bar_start_row}'] = "Transaksi Per Bulan"
            ws_chart[f'A{bar_start_row}'].style = styles['subtitle']
            
            # Header untuk bar chart data
            bar_headers = ['Bulan', 'Jumlah Transaksi']
            for col, header in enumerate(bar_headers, 1):
                cell = ws_chart.cell(row=bar_start_row + 1, column=col, value=header)
                cell.style = styles['header']
            
            # Data bar chart
            row = bar_start_row + 2
            for month in sorted(monthly_data.keys()):
                month_formatted = datetime.strptime(month, '%Y-%m').strftime('%b %Y')
                ws_chart[f'A{row}'] = month_formatted
                ws_chart[f'B{row}'] = monthly_data[month]
                ws_chart[f'A{row}'].style = styles['data']
                ws_chart[f'B{row}'].style = styles['center']
                row += 1
            
            # Buat Bar Chart
            bar_chart = BarChart()
            bar_chart.title = "Trend Transaksi Per Bulan"
            bar_chart.height = 10
            bar_chart.width = 15
            bar_chart.x_axis.title = "Bulan"
            bar_chart.y_axis.title = "Jumlah Transaksi"
            
            data_range = Reference(ws_chart, min_col=2, min_row=bar_start_row + 1,
                                 max_row=bar_start_row + 1 + len(monthly_data))
            categories = Reference(ws_chart, min_col=1, min_row=bar_start_row + 2,
                                 max_row=bar_start_row + 1 + len(monthly_data))
            
            bar_chart.add_data(data_range, titles_from_data=True)
            bar_chart.set_categories(categories)
            
            ws_chart.add_chart(bar_chart, f'D{bar_start_row}')
        
        # Auto-adjust column width
        for col in range(1, 9):
            ws_chart.column_dimensions[get_column_letter(col)].width = 15
        
        return ws_chart
    
    def _create_data_sheet(self, wb, transaksi_list, sheet_name="Data Transaksi"):
        """Buat sheet dengan data transaksi"""
        ws = wb.active
        ws.title = sheet_name
        styles = self._create_styles()
        
        # Title
        ws['A1'] = "DATA TRANSAKSI LENGKAP"
        ws['A1'].style = styles['title']
        ws.merge_cells('A1:H1')
        
        # Info export
        ws['A2'] = f"Export Date: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Header kolom
        headers = [
            'No',
            'Nomor Invoice',
            'Nama Pembeli',
            'Tanggal',
            'Subtotal (Rp)',
            'Pembayaran (Rp)',
            'Sisa Piutang (Rp)',
            'Status',
            'Rincian Barang'
        ]
        
        # Tulis header
        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.style = styles['header']
        
        # Tulis data transaksi
        row = header_row + 1
        for idx, transaksi in enumerate(transaksi_list, 1):
            transaksi_id = transaksi[0]
            nama_pembeli = transaksi[1]
            subtotal = transaksi[2]
            pembayaran = transaksi[3]
            status = transaksi[4]
            tanggal = transaksi[5]
            no_invoice = transaksi[6]
            
            # Hitung sisa piutang
            sisa_piutang = subtotal - pembayaran
            
            # Format tanggal
            try:
                tanggal_formatted = datetime.strptime(tanggal, '%Y-%m-%d').strftime('%d-%m-%Y')
            except:
                tanggal_formatted = tanggal
            
            # Ambil detail barang untuk transaksi ini
            detail_barang = self.db.get_transaksi_detail(transaksi_id)
            rincian_text = ""
            
            for detail in detail_barang:
                warna = detail[9]
                kategori = detail[10]
                ukuran = detail[11]
                qty = detail[7]
                harga = detail[8]
                
                rincian_text += f"• {warna} - {kategori} - {ukuran} (Qty: {qty}, @{self.format_rupiah(harga)})\n"
            
            # Tulis data ke Excel
            ws.cell(row=row, column=1, value=idx).style = styles['center']
            ws.cell(row=row, column=2, value=no_invoice).style = styles['data']
            ws.cell(row=row, column=3, value=nama_pembeli).style = styles['data']
            ws.cell(row=row, column=4, value=tanggal_formatted).style = styles['center']
            ws.cell(row=row, column=5, value=self.format_rupiah(subtotal)).style = styles['number']
            ws.cell(row=row, column=6, value=self.format_rupiah(pembayaran)).style = styles['number']
            ws.cell(row=row, column=7, value=self.format_rupiah(sisa_piutang)).style = styles['number']
            ws.cell(row=row, column=8, value=status).style = styles['center']
            ws.cell(row=row, column=9, value=rincian_text.strip()).style = styles['data']
            
            # Set warna untuk status
            status_cell = ws.cell(row=row, column=8)
            if status == 'LUNAS':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Set row height untuk menampung rincian barang
            ws.row_dimensions[row].height = max(60, len(detail_barang) * 15)
            
            row += 1
        
        # Auto-adjust column width
        column_widths = [5, 18, 20, 12, 15, 15, 15, 12, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        return ws
    
    def export_transaksi(self):
        """Export semua transaksi ke Excel dengan multiple sheets"""
        try:
            # Ambil semua data transaksi
            transaksi_list = self.db.get_all_transaksi()
            
            if not transaksi_list:
                return False, "Tidak ada data transaksi untuk diexport"
            
            # Buat workbook baru
            wb = openpyxl.Workbook()
            
            # Buat sheets
            self._create_data_sheet(wb, transaksi_list)
            self._create_summary_sheet(wb, transaksi_list)
            self._create_chart_sheet(wb, transaksi_list)
            
            # Set active sheet ke Data Transaksi
            wb.active = wb["Data Transaksi"]
            
            # Simpan file
            filename = f"laporan_transaksi_lengkap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            return True, f"Laporan berhasil diexport ke: {filepath}"
            
        except Exception as e:
            return False, f"Error saat export: {str(e)}"
    
    def export_transaksi_by_date(self, start_date, end_date):
        """Export transaksi berdasarkan range tanggal dengan multiple sheets"""
        try:
            # Query transaksi berdasarkan tanggal
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT id, nama_pembeli, subtotal, pembayaran, status, tanggal, no_invoice
                FROM transaksi 
                WHERE tanggal BETWEEN ? AND ?
                ORDER BY id DESC
            ''', (start_date, end_date))
            
            transaksi_list = cursor.fetchall()
            conn.close()
            
            if not transaksi_list:
                return False, f"Tidak ada transaksi pada periode {start_date} sampai {end_date}"
            
            # Buat workbook baru
            wb = openpyxl.Workbook()
            
            # Period info for titles
            period_info = f"PERIODE {start_date} s/d {end_date}"
            
            # Buat sheets
            self._create_data_sheet(wb, transaksi_list, f"Data_Transaksi_{start_date}_to_{end_date}")
            self._create_summary_sheet(wb, transaksi_list, period_info)
            self._create_chart_sheet(wb, transaksi_list)
            
            # Set active sheet ke Data Transaksi
            wb.active = wb[f"Data_Transaksi_{start_date}_to_{end_date}"]
            
            # Simpan file
            filename = f"laporan_transaksi_{start_date}_to_{end_date}_{datetime.now().strftime('%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            return True, f"Laporan periode {start_date} s/d {end_date} berhasil diexport ke: {filepath}"
            
        except Exception as e:
            return False, f"Error saat export: {str(e)}"
    
    def export_outstanding_payments(self):
        """Export khusus untuk transaksi yang belum lunas"""
        try:
            # Ambil transaksi yang belum lunas
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT id, nama_pembeli, subtotal, pembayaran, status, tanggal, no_invoice
                FROM transaksi 
                WHERE status = 'BELUM LUNAS'
                ORDER BY tanggal ASC
            ''')
            
            transaksi_list = cursor.fetchall()
            conn.close()
            
            if not transaksi_list:
                return False, "Tidak ada transaksi yang belum lunas"
            
            # Buat workbook baru
            wb = openpyxl.Workbook()
            
            # Buat sheets
            self._create_data_sheet(wb, transaksi_list, "Outstanding_Payments")
            self._create_summary_sheet(wb, transaksi_list, "TRANSAKSI BELUM LUNAS")
            
            # Hapus chart sheet karena tidak relevan untuk outstanding payments
            if "Charts" in wb.sheetnames:
                wb.remove(wb["Charts"])
            
            # Set active sheet
            wb.active = wb["Outstanding_Payments"]
            
            # Simpan file
            filename = f"outstanding_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)
            
            return True, f"Laporan piutang berhasil diexport ke: {filepath}"
            
        except Exception as e:
            return False, f"Error saat export: {str(e)}"

class ExcelExporterGudang:
    """Kelas terpisah untuk export data gudang ke Excel."""

    def __init__(self, db_gudang):
        self.db_gudang = db_gudang
        self.export_folder = 'export'
        self._ensure_export_folder()

    def _ensure_export_folder(self):
        if not os.path.exists(self.export_folder):
            os.makedirs(self.export_folder)

    def _create_styles(self):
        """Buat style Excel — sama dengan ExcelExporter utama."""
        styles = {}

        def make_style(name, font_kw, fill_color=None, align_h="left", number_fmt=None):
            s = NamedStyle(name=name)
            s.font = Font(**font_kw)
            if fill_color:
                s.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            s.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)
            s.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if number_fmt:
                s.number_format = number_fmt
            return s

        styles['header'] = make_style('gudang_header', {'bold': True, 'color': 'FFFFFF', 'size': 11},
                                      fill_color='2F4F4F', align_h='center')
        styles['data'] = make_style('gudang_data', {'size': 10})
        styles['center'] = make_style('gudang_center', {'size': 10}, align_h='center')
        styles['number'] = make_style('gudang_number', {'size': 10}, align_h='right')
        styles['title'] = NamedStyle(name='gudang_title')
        styles['title'].font = Font(bold=True, size=16, color='2F4F4F')
        styles['title'].alignment = Alignment(horizontal='center')
        styles['subtitle'] = NamedStyle(name='gudang_subtitle')
        styles['subtitle'].font = Font(bold=True, size=12, color='4F4F4F')
        styles['ok'] = make_style('gudang_ok', {'size': 10}, fill_color='C6EFCE', align_h='center')
        styles['menipis'] = make_style('gudang_menipis', {'size': 10}, fill_color='FFEB9C', align_h='center')
        styles['habis'] = make_style('gudang_habis', {'bold': True, 'size': 10}, fill_color='FFC7CE', align_h='center')
        return styles

    def _sheet_stok(self, wb, data_stok):
        """Sheet utama: data stok barang."""
        ws = wb.active
        ws.title = "Data Stok"
        styles = self._create_styles()

        ws['A1'] = "LAPORAN DATA STOK GUDANG"
        ws['A1'].style = styles['title']
        ws.merge_cells('A1:H1')

        ws['A2'] = f"Export: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)

        headers = ['No', 'ID', 'Kategori', 'Warna', 'Ukuran', 'Jumlah Stok', 'Stok Minimum', 'Update Terakhir', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.style = styles['header']

        for idx, item in enumerate(data_stok, 1):
            row = 4 + idx
            jumlah = item['jumlah']
            minimum = item['stok_minimum']
            if jumlah == 0:
                status, st_key = "HABIS", 'habis'
            elif jumlah <= minimum:
                status, st_key = "MENIPIS", 'menipis'
            else:
                status, st_key = "OK", 'ok'

            vals = [idx, item['id'], item['kategori'], item['warna'], item['ukuran'],
                    jumlah, minimum, item['tanggal_update']]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col in [1, 2, 6, 7]:
                    cell.style = styles['center']
                else:
                    cell.style = styles['data']

            status_cell = ws.cell(row=row, column=9, value=status)
            status_cell.style = styles[st_key]

        widths = [5, 6, 18, 16, 12, 14, 14, 20, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        return ws

    def _sheet_summary_stok(self, wb, data_stok):
        """Sheet summary statistik gudang."""
        ws = wb.create_sheet("Summary Gudang")
        styles = self._create_styles()

        ws['A1'] = "SUMMARY STATISTIK GUDANG"
        ws['A1'].style = styles['title']
        ws.merge_cells('A1:D1')

        ws['A3'] = "Tanggal Export:"
        ws['B3'] = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        total_item = len(data_stok)
        total_stok = sum(d['jumlah'] for d in data_stok)
        total_habis = sum(1 for d in data_stok if d['jumlah'] == 0)
        total_menipis = sum(1 for d in data_stok if 0 < d['jumlah'] <= d['stok_minimum'])
        total_ok = total_item - total_habis - total_menipis

        summary_rows = [
            ("STATISTIK UMUM", ""),
            ("Total Jenis Barang", total_item),
            ("Total Unit Stok", total_stok),
            ("Stok OK", total_ok),
            ("Stok Menipis", total_menipis),
            ("Stok Habis", total_habis),
        ]

        # Per Kategori
        from collections import defaultdict
        kat_stats = defaultdict(lambda: {'jumlah': 0, 'item': 0})
        for d in data_stok:
            kat_stats[d['kategori']]['jumlah'] += d['jumlah']
            kat_stats[d['kategori']]['item'] += 1

        start = 5
        for i, (label, val) in enumerate(summary_rows):
            r = start + i
            ws[f'A{r}'] = label
            if label.isupper():
                ws[f'A{r}'].style = styles['subtitle']
                ws.merge_cells(f'A{r}:D{r}')
            else:
                ws[f'A{r}'].style = styles['data']
                ws[f'B{r}'] = val
                ws[f'B{r}'].style = styles['center']

        kat_start = start + len(summary_rows) + 2
        ws[f'A{kat_start}'] = "STOK PER KATEGORI"
        ws[f'A{kat_start}'].style = styles['subtitle']
        ws.merge_cells(f'A{kat_start}:D{kat_start}')

        h_row = kat_start + 1
        for col, h in enumerate(['Kategori', 'Jumlah Item', 'Total Unit Stok'], 1):
            ws.cell(row=h_row, column=col, value=h).style = styles['header']

        for i, (kat, st) in enumerate(sorted(kat_stats.items())):
            r = h_row + 1 + i
            ws.cell(row=r, column=1, value=kat).style = styles['data']
            ws.cell(row=r, column=2, value=st['item']).style = styles['center']
            ws.cell(row=r, column=3, value=st['jumlah']).style = styles['center']

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 22

        return ws

    def _sheet_log(self, wb, data_log):
        """Sheet log perubahan stok."""
        ws = wb.create_sheet("Log Stok")
        styles = self._create_styles()

        ws['A1'] = "LOG PERUBAHAN STOK"
        ws['A1'].style = styles['title']
        ws.merge_cells('A1:I1')

        headers = ['No', 'ID', 'Warna', 'Kategori', 'Ukuran', 'Stok Lama', 'Stok Baru', 'Perubahan', 'Aksi', 'Tanggal']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).style = styles['header']

        for idx, item in enumerate(data_log, 1):
            row = 3 + idx
            perubahan = item['perubahan']
            perubahan_str = f"+{perubahan}" if perubahan > 0 else str(perubahan)
            vals = [idx, item['id'], item['warna'], item['kategori'], item['ukuran'],
                    item['stok_lama'], item['stok_baru'], perubahan_str, item['aksi'], item['tanggal']]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.style = styles['center'] if col in [1, 2, 6, 7, 8] else styles['data']
                if col == 8:
                    from openpyxl.styles import Color
                    if perubahan > 0:
                        cell.font = Font(size=10, color='375623')
                    elif perubahan < 0:
                        cell.font = Font(size=10, color='9C0006')

        for i, w in enumerate([5, 6, 14, 16, 12, 12, 12, 12, 14, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    def _sheet_backup(self, wb, data_backup):
        """Sheet backup stok."""
        ws = wb.create_sheet("Backup Stok")
        styles = self._create_styles()

        ws['A1'] = "BACKUP STOK"
        ws['A1'].style = styles['title']
        ws.merge_cells('A1:G1')

        headers = ['No', 'ID', 'Warna', 'Kategori', 'Ukuran', 'Jumlah', 'Keterangan', 'Tanggal Backup']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).style = styles['header']

        for idx, item in enumerate(data_backup, 1):
            row = 3 + idx
            vals = [idx, item['id'], item['warna'], item['kategori'], item['ukuran'],
                    item['jumlah'], item.get('keterangan', ''), item['tanggal_backup']]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.style = styles['center'] if col in [1, 2, 6] else styles['data']

        for i, w in enumerate([5, 6, 14, 16, 12, 10, 24, 22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    def export_gudang(self):
        """Export semua data gudang ke Excel dengan multiple sheets."""
        try:
            data_stok = self.db_gudang.get_all_stok()
            data_log = self.db_gudang.get_log_stok(limit=9999)
            data_backup = self.db_gudang.get_backup_stok(limit=9999)

            if not data_stok:
                return False, "Tidak ada data stok untuk diexport"

            wb = openpyxl.Workbook()

            self._sheet_stok(wb, data_stok)
            self._sheet_summary_stok(wb, data_stok)
            self._sheet_log(wb, data_log)
            self._sheet_backup(wb, data_backup)

            wb.active = wb["Data Stok"]

            filename = f"laporan_gudang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            wb.save(filepath)

            return True, f"Laporan gudang berhasil diexport ke: {filepath}"

        except Exception as e:
            return False, f"Error saat export gudang: {str(e)}"
